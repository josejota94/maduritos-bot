import json
import math
import os
import threading
import unicodedata

import psycopg2

# ---------------------------------------------------------------------------
# CONFIGURACIÓN DEL NEGOCIO (edita estos valores a tu gusto, o mejor en .env)
# ---------------------------------------------------------------------------
LAT_LOCAL = float(os.getenv("LAT_LOCAL", "-6.0350"))
LON_LOCAL = float(os.getenv("LON_LOCAL", "-76.9700"))
RADIO_MAXIMO_KM = float(os.getenv("RADIO_MAXIMO_KM", "4.0"))
def _entero_env(nombre_var: str, valor_por_defecto: int) -> int:
    """Lee una variable de entorno como número entero, tolerando que alguien
    haya puesto sin querer un valor con decimales (ej. '3.5') — lo redondea
    en vez de tumbar el servidor al arrancar."""
    crudo = os.getenv(nombre_var, str(valor_por_defecto))
    try:
        return int(float(crudo))
    except (TypeError, ValueError):
        return valor_por_defecto


PEDIDO_MINIMO_UNIDADES = _entero_env("PEDIDO_MINIMO_UNIDADES", 3)
DIRECCION_LOCAL = os.getenv("DIRECCION_LOCAL", "Jr. Lima cuadra 4, frente al Banco BCP")

PRECIOS = {"mediano": 3.00, "grande": 4.00, "relleno_extra": 5.00}
TAMANOS_TITULOS = {"mediano": "Mediano", "grande": "Grande", "relleno_extra": "Relleno Extra"}

# Sabores base y TODAS las combinaciones posibles entre ellos (dobles + la triple).
# El precio es el mismo de siempre (según tamaño), combinar sabores no tiene costo extra.
_SABORES_BASE = ["queso", "mani", "chicharron"]
RELLENOS = [
    "queso", "mani", "chicharron",
    "queso_mani", "queso_chicharron", "mani_chicharron",
    "combinado",
]
RELLENOS_TITULOS = {
    "queso": "Queso",
    "mani": "Maní",
    "chicharron": "Chicharrón",
    "queso_mani": "Queso + Maní",
    "queso_chicharron": "Queso + Chicharrón",
    "mani_chicharron": "Maní + Chicharrón",
    "combinado": "Combinado (Queso + Maní + Chicharrón)",
}

# ---------------------------------------------------------------------------
# BASE DE DATOS: Postgres, en vez del archivo SQLite de antes.
#
# El archivo SQLite (pedidos_maduritos.db) vivía SOLO dentro del contenedor
# de Render. Como el plan free no tiene disco persistente y el servicio se
# duerme/reinicia (inactividad, falta de memoria, redeploy), ese archivo se
# borraba por completo en cada reinicio — junto con TODAS las conversaciones
# en curso. Por eso el bot "se olvidaba" a mitad de un pedido y volvía a
# mandar el saludo desde cero. Postgres vive fuera del contenedor, así que
# sobrevive a los reinicios.
#
# DATABASE_URL debe apuntar a esa base (ver render.yaml).
# ---------------------------------------------------------------------------
DATABASE_URL = os.getenv("DATABASE_URL", "")
if not DATABASE_URL:
    raise RuntimeError(
        "Falta la variable de entorno DATABASE_URL — debe apuntar a tu base "
        "de datos Postgres (revisa la configuración en Render)."
    )

if "sslmode=" not in DATABASE_URL:
    _separador = "&" if "?" in DATABASE_URL else "?"
    DATABASE_URL = f"{DATABASE_URL}{_separador}sslmode=require"

_lock = threading.Lock()
conn = psycopg2.connect(DATABASE_URL)
conn.autocommit = False
cursor = conn.cursor()

with _lock:
    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS pedidos (
            id SERIAL PRIMARY KEY,
            telefono TEXT,
            cliente_nombre TEXT,
            detalle TEXT,
            total_unidades INTEGER,
            monto_total REAL,
            latitud REAL,
            longitud REAL,
            distancia_km REAL,
            estado TEXT DEFAULT 'PENDIENTE',
            fecha TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
        """
    )
    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS sesiones (
            telefono TEXT PRIMARY KEY,
            nombre TEXT,
            paso TEXT,
            cantidad_total INTEGER,
            unidad_actual INTEGER DEFAULT 0,
            tamano_pendiente TEXT,
            carrito TEXT
        )
        """
    )
    # Migración suave por si la tabla ya existía de una versión anterior
    cursor.execute("ALTER TABLE sesiones ADD COLUMN IF NOT EXISTS unidad_actual INTEGER DEFAULT 0")
    cursor.execute("ALTER TABLE sesiones ADD COLUMN IF NOT EXISTS tamano_pendiente TEXT")

    # Tabla que main.py ya esperaba (bl.guardar_suscripcion_push /
    # bl.eliminar_suscripcion_push) pero que no existía todavía.
    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS suscripciones_push (
            endpoint TEXT PRIMARY KEY,
            p256dh TEXT,
            auth TEXT
        )
        """
    )
    conn.commit()


# ---------------------------------------------------------------------------
# UTILIDADES
# ---------------------------------------------------------------------------
def _normalizar(texto: str) -> str:
    texto = (texto or "").strip().lower()
    texto = "".join(
        c for c in unicodedata.normalize("NFD", texto) if unicodedata.category(c) != "Mn"
    )
    return texto


def calcular_distancia(lat1, lon1, lat2, lon2):
    R = 6371.0
    dlat = math.radians(lat2 - lat1)
    dlon = math.radians(lon2 - lon1)
    a = (
        math.sin(dlat / 2) ** 2
        + math.cos(math.radians(lat1)) * math.cos(math.radians(lat2)) * math.sin(dlon / 2) ** 2
    )
    return R * (2 * math.atan2(math.sqrt(a), math.sqrt(1 - a)))


def _texto(msg):
    return {"tipo": "texto", "texto": msg}


def _botones(msg, opciones):
    """opciones: lista de (id, titulo) — máximo 3, límite real de WhatsApp."""
    return {"tipo": "botones", "texto": msg, "opciones": [{"id": i, "titulo": t} for i, t in opciones]}


def _lista(msg, opciones, boton_texto="Elegir"):
    """opciones: lista de (id, titulo) — hasta 10, usa el formato de lista de WhatsApp."""
    return {"tipo": "lista", "texto": msg, "boton_texto": boton_texto, "opciones": [{"id": i, "titulo": t} for i, t in opciones]}


def _valor_opcion(texto_o_id: str, prefijo: str) -> str:
    """
    Normaliza tanto un id de botón/lista (ej. 'tam_grande') como texto libre
    escrito a mano (ej. 'Grande') a un mismo valor comparable ('grande').
    """
    v = _normalizar(texto_o_id)
    if v.startswith(prefijo):
        v = v[len(prefijo):]
    return v


def _detectar_tamano(texto_o_id: str):
    """
    Reconoce el tamaño elegido, ya sea un id de botón (ej. 'tam_relleno_extra')
    o texto libre escrito a mano (ej. 'Relleno Extra', con espacio en vez de
    guion bajo). Devuelve la key de PRECIOS o None si no se reconoce nada.
    """
    v = _valor_opcion(texto_o_id or "", "tam_")
    if v in PRECIOS:
        return v
    v_guion = v.replace(" ", "_")
    if v_guion in PRECIOS:
        return v_guion
    return None


_CONFIRMAR_PALABRAS = {"confirmar_pedido", "confirmar", "confirmo", "si", "sí", "s", "ok", "okay", "dale", "claro", "yes", "correcto", "va"}
_CANCELAR_PALABRAS = {"cancelar_pedido", "cancelar", "cancelo", "no", "n", "cancela"}


def _resolver_por_posicion(texto_o_id: str, opciones):
    """
    Si el cliente responde con un número suelto (ej. '2') que coincide con
    la POSICIÓN de una de las opciones que se le mostraron (ej. la 2da de la
    lista), lo traduce al id real de esa opción (ej. 'tam_grande').
    Necesario porque, por WhatsApp-QR, las opciones a veces solo se pueden
    mostrar como texto numerado ("1. Mediano / 2. Grande / ..."), y sin esto
    el cliente escribe "2" y el bot no lo reconoce — repite la pregunta en
    bucle en vez de avanzar.
    """
    crudo = (texto_o_id or "").strip()
    if crudo.isdigit():
        idx = int(crudo) - 1
        if 0 <= idx < len(opciones):
            return opciones[idx][0]
    return texto_o_id


def _detectar_confirmacion(texto_o_id: str):
    """Devuelve 'confirmar', 'cancelar' o None. Acepta el id del botón o
    sinónimos comunes escritos a mano ('si', 'sí', 'ok', 'no', ...)."""
    v = _normalizar(texto_o_id or "")
    if v in _CANCELAR_PALABRAS:
        return "cancelar"
    if v in _CONFIRMAR_PALABRAS:
        return "confirmar"
    return None


def _detectar_relleno(texto_o_id: str):
    """
    Reconoce el relleno elegido, ya sea:
    - un id de la lista de WhatsApp (ej. 'rel_queso_mani' -> 'queso_mani'), o
    - texto libre escrito a mano (ej. 'queso y mani', 'los 3', 'combinado', 'todo').
    Devuelve el valor canónico (una key de RELLENOS) o None si no se reconoce nada.
    """
    v = _valor_opcion(texto_o_id or "", "rel_")
    if v in RELLENOS:
        return v

    if v in ("combinado", "todo", "todos", "los3", "los 3", "de todo"):
        return "combinado"

    presentes = [s for s in _SABORES_BASE if s in v or (s == "mani" and "man" in v)]
    if len(presentes) >= 3:
        return "combinado"
    if len(presentes) == 2:
        # Orden fijo para que coincida con las keys de RELLENOS_TITULOS
        presentes.sort(key=_SABORES_BASE.index)
        return "_".join(presentes)
    if len(presentes) == 1:
        return presentes[0]
    return None


def _get_sesion(telefono):
    with _lock:
        cursor.execute(
            "SELECT telefono, nombre, paso, cantidad_total, unidad_actual, tamano_pendiente, carrito "
            "FROM sesiones WHERE telefono=%s",
            (telefono,),
        )
        row = cursor.fetchone()
    if not row:
        return None
    return {
        "telefono": row[0],
        "nombre": row[1],
        "paso": row[2],
        "cantidad_total": row[3],
        "unidad_actual": row[4] or 0,
        "tamano_pendiente": row[5],
        "carrito": json.loads(row[6]) if row[6] else [],
    }


def _guardar_sesion(telefono, nombre, paso, cantidad_total=None, unidad_actual=0, tamano_pendiente=None, carrito=None):
    with _lock:
        cursor.execute(
            """
            INSERT INTO sesiones (telefono, nombre, paso, cantidad_total, unidad_actual, tamano_pendiente, carrito)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
            ON CONFLICT(telefono) DO UPDATE SET
                nombre=excluded.nombre,
                paso=excluded.paso,
                cantidad_total=excluded.cantidad_total,
                unidad_actual=excluded.unidad_actual,
                tamano_pendiente=excluded.tamano_pendiente,
                carrito=excluded.carrito
            """,
            (telefono, nombre, paso, cantidad_total, unidad_actual, tamano_pendiente, json.dumps(carrito or [])),
        )
        conn.commit()


def _borrar_sesion(telefono):
    with _lock:
        cursor.execute("DELETE FROM sesiones WHERE telefono=%s", (telefono,))
        conn.commit()


def _opciones_tamano():
    return [
        ("tam_mediano", "Mediano (S/ 3.00)"),
        ("tam_grande", "Grande (S/ 4.00)"),
        ("tam_relleno_extra", "Relleno Extra (S/ 5.00)"),
    ]


def _opciones_relleno():
    return [(f"rel_{r}", RELLENOS_TITULOS[r]) for r in RELLENOS]


def _pedir_tamano(unidad_actual, cantidad_total):
    return _botones(
        f"Maduro *{unidad_actual}/{cantidad_total}* — elige el tamaño:",
        _opciones_tamano(),
    )


def _pedir_relleno(unidad_actual, cantidad_total, tamano):
    nombre_tam = TAMANOS_TITULOS.get(tamano, tamano)
    return _lista(
        f"Maduro {unidad_actual}/{cantidad_total} ({nombre_tam}) — elige el relleno:",
        _opciones_relleno(),
        boton_texto="Elegir relleno",
    )


def _resumen_carrito(carrito):
    return "\n".join(
        f"- {TAMANOS_TITULOS.get(c['tamano'], c['tamano'])} {RELLENOS_TITULOS[c['relleno']]} (S/ {c['precio']:.2f})"
        for c in carrito
    )


def _detalle_agrupado(carrito, con_precios=True):
    """
    Junta los maduritos iguales (mismo tamaño + relleno) y devuelve un texto
    de varias líneas, tipo lista de cocina/despacho, ej.:
        2x Grande - Queso + Chicharrón — S/ 8.00
        1x Relleno Extra - Combinado (Queso + Maní + Chicharrón) — S/ 5.00
    Mucho más claro para el repartidor y para pedirle a la cocinera que
    "grande queso_chicharron, grande queso_chicharron, relleno_extra combinado".
    Con `con_precios=True` (para el repartidor) cada línea trae su subtotal,
    así sabe cuánto cobrar por cada grupo aunque el cliente pague por partes.
    """
    conteo = {}
    orden = []
    for c in carrito:
        clave = (c["tamano"], c["relleno"])
        if clave not in conteo:
            conteo[clave] = {"cantidad": 0, "precio_unit": c["precio"]}
            orden.append(clave)
        conteo[clave]["cantidad"] += 1
    lineas = []
    for tamano, relleno in orden:
        info = conteo[(tamano, relleno)]
        cantidad = info["cantidad"]
        nombre_tam = TAMANOS_TITULOS.get(tamano, tamano)
        nombre_rel = RELLENOS_TITULOS.get(relleno, relleno)
        linea = f"{cantidad}x {nombre_tam} - {nombre_rel}"
        if con_precios:
            subtotal = cantidad * info["precio_unit"]
            linea += f" — S/ {subtotal:.2f}"
        lineas.append(linea)
    return "\n".join(lineas)


_AGRADECIMIENTO_PALABRAS = {
    "gracias", "muchas gracias", "grac", "gracias!", "muchisimas gracias",
    "ok gracias", "genial", "perfecto", "excelente", "buenisimo", "de nada",
}


def _es_agradecimiento(texto: str) -> bool:
    v = _normalizar(texto or "")
    if v in _AGRADECIMIENTO_PALABRAS:
        return True
    return "gracias" in v and len(v) <= 40


MENU_INTRO = (
    "¡Hola{saludo}! Bienvenido a *Maduritos Asados* 🍌🔥\n\n"
    "*Carta:*\n"
    "- Mediano: S/ 3.00\n"
    "- Grande: S/ 4.00\n"
    "- Relleno Extra: S/ 5.00 🔥\n"
    "_Sabores: Queso, Maní, Chicharrón — o cualquier combinación entre ellos "
    "(2 sabores, o los 3 juntos), sin costo extra._\n\n"
    f"*Condición:* pedido mínimo de {PEDIDO_MINIMO_UNIDADES} unidades para delivery.\n\n"
    "_Escribe *cancelar* en cualquier momento para reiniciar tu pedido._"
)


def _opciones_cantidad():
    return [("cant_3", "3 unidades"), ("cant_5", "5 unidades"), ("cant_otra", "Otra cantidad")]


def _pedir_cantidad(nombre):
    msg = MENU_INTRO.format(saludo=f" {nombre}" if nombre else "") + "\n\n¿Cuántos maduritos en total deseas pedir?"
    return _botones(msg, _opciones_cantidad())


def procesar_mensaje(telefono: str, nombre: str, tipo: str, texto: str = None, lat: float = None, lon: float = None):
    """
    Motor del flujo conversacional. Usado tanto por el webhook real de WhatsApp
    como por el simulador de pruebas (/simulador).

    Devuelve un diccionario describiendo la respuesta:
      {"tipo": "texto", "texto": "..."}
      {"tipo": "botones", "texto": "...", "opciones": [{"id":..,"titulo":..}, ...]}   (máx. 3)
      {"tipo": "lista",   "texto": "...", "boton_texto": "...", "opciones": [...]}     (hasta 10)

    Tanto un botón/lista real de WhatsApp como texto escrito a mano llegan aquí
    como tipo="texto" con `texto` = id de la opción (ej. "tam_grande") o palabra libre
    (ej. "grande") — ambos se aceptan.
    """
    sesion = _get_sesion(telefono)

    if texto and _normalizar(texto) in ("cancelar", "reiniciar"):
        _borrar_sesion(telefono)
        return _texto("Tu pedido fue cancelado. Escribe cualquier mensaje para empezar de nuevo. 🙂")

    # --- Sin sesión activa ---
    if sesion is None:
        # Si el cliente solo está agradeciendo/despidiéndose después de un
        # pedido ya confirmado, no lo metemos de nuevo al flujo de "cuántos
        # maduritos quieres" — respondemos amable y quedamos ahí.
        if texto and _es_agradecimiento(texto):
            return _texto(
                "¡Con gusto! 🍌 Tu pedido ya fue tomado y está en camino. "
                "Cualquier consulta, escríbenos por aquí."
            )
        _guardar_sesion(telefono, nombre, "esperando_cantidad")
        return _pedir_cantidad(nombre)

    paso = sesion["paso"]

    # --- Paso 1: cantidad total ---
    if paso == "esperando_cantidad":
        crudo = (texto or "").strip()
        valor = _valor_opcion(crudo, "cant_")
        opciones = _opciones_cantidad()

        if valor == "otra":
            return _texto("Escribe el número total de maduritos que deseas pedir.")

        cantidad = None
        if valor.isdigit():
            cantidad = int(valor)
        elif crudo.isdigit():
            cantidad = int(crudo)
        else:
            return _pedir_cantidad(nombre)

        # Corrección de ambigüedad: si un canal le muestra al cliente las
        # opciones numeradas en texto plano ("1. 3 unidades / 2. 5 unidades /
        # 3. Otra cantidad") y el cliente responde solo "1" pensando "elijo la
        # opción 1", eso llega aquí como el número 1 suelto. Como una cantidad
        # así de chica (menor al mínimo) de todas formas sería inválida por sí
        # sola, y coincide con la posición de una opción ofrecida, asumimos que
        # quiso decir "esa opción" en vez de rechazarlo como pedido inválido.
        if cantidad < PEDIDO_MINIMO_UNIDADES and 1 <= cantidad <= len(opciones):
            id_opcion = opciones[cantidad - 1][0]
            if id_opcion == "cant_otra":
                return _texto("Escribe el número total de maduritos que deseas pedir.")
            valor_opcion = _valor_opcion(id_opcion, "cant_")
            if valor_opcion.isdigit():
                cantidad = int(valor_opcion)

        if cantidad < PEDIDO_MINIMO_UNIDADES:
            resp = _pedir_cantidad(nombre)
            resp["texto"] = (
                f"El pedido mínimo para delivery es de {PEDIDO_MINIMO_UNIDADES} unidades. "
                f"¿Cuántos maduritos deseas pedir?"
            )
            return resp

        # Antes de armar el pedido, confirmamos el nombre del cliente (para que
        # el repartidor sepa a quién llamar y a quién entregarle) — usamos el
        # número de WhatsApp de siempre como número de contacto, no hace falta
        # pedirlo aparte.
        _guardar_sesion(telefono, nombre, "esperando_nombre", cantidad_total=cantidad, unidad_actual=1, carrito=[])
        return _texto("¿A nombre de quién anotamos el pedido? ")

    # --- Paso 1b: nombre del cliente ---
    if paso == "esperando_nombre":
        nombre_cliente = (texto or "").strip()
        if not nombre_cliente:
            return _texto("¿A nombre de quién anotamos el pedido?")
        
        # Guardamos el nombre y pasamos al nuevo paso del celular
        _guardar_sesion(
            telefono, nombre_cliente, "esperando_contacto",
            cantidad_total=sesion["cantidad_total"], unidad_actual=1, carrito=[],
        )
        return _texto(f"¡Anotado, {nombre_cliente}! 📱 ¿A qué número de celular puede llamarte o escribirte el repartidor al llegar?")

    # --- NUEVO Paso 1c: número de contacto ---
    if paso == "esperando_contacto":
        crudo = (texto or "").strip()
        
        # Filtramos para quedarnos ÚNICAMENTE con los números (quita guiones, letras y espacios)
        numero_limpio = "".join(c for c in crudo if c.isdigit())
        
        # Validar formato de Perú (empieza con 9 y tiene 9 dígitos exactos)
        if len(numero_limpio) != 9 or not numero_limpio.startswith("9"):
            return _texto("Ese número no parece válido. Por favor, asegúrate de escribir tu celular completo de 9 dígitos (ejemplo: 987654321).")
        
        # Juntamos el nombre y el celular validado
        nombre_con_celular = f"{sesion['nombre']} | {numero_limpio}"
        
        _guardar_sesion(
            telefono, nombre_con_celular, "esperando_tamano",
            cantidad_total=sesion["cantidad_total"], unidad_actual=1, carrito=[],
        )
        return _pedir_tamano(1, sesion["cantidad_total"])

    # --- Paso 2a: tamaño del maduro actual ---
    if paso == "esperando_tamano":
        valor = _detectar_tamano(_resolver_por_posicion(texto, _opciones_tamano()))
        if not valor:
            resp = _pedir_tamano(sesion["unidad_actual"], sesion["cantidad_total"])
            resp["texto"] = "No entendí esa opción. " + resp["texto"]
            return resp
        _guardar_sesion(
            telefono, sesion["nombre"], "esperando_relleno",
            cantidad_total=sesion["cantidad_total"], unidad_actual=sesion["unidad_actual"],
            tamano_pendiente=valor, carrito=sesion["carrito"],
        )
        return _pedir_relleno(sesion["unidad_actual"], sesion["cantidad_total"], valor)

    # --- Paso 2b: relleno del maduro actual ---
    if paso == "esperando_relleno":
        valor = _detectar_relleno(_resolver_por_posicion(texto, _opciones_relleno()))
        if not valor:
            resp = _pedir_relleno(sesion["unidad_actual"], sesion["cantidad_total"], sesion["tamano_pendiente"])
            resp["texto"] = "No entendí ese relleno. " + resp["texto"]
            return resp

        tamano = sesion["tamano_pendiente"]
        carrito = sesion["carrito"] + [{"tamano": tamano, "relleno": valor, "precio": PRECIOS[tamano]}]
        siguiente_unidad = sesion["unidad_actual"] + 1
        cantidad_total = sesion["cantidad_total"]

        if siguiente_unidad <= cantidad_total:
            _guardar_sesion(
                telefono, sesion["nombre"], "esperando_tamano",
                cantidad_total=cantidad_total, unidad_actual=siguiente_unidad,
                carrito=carrito,
            )
            return _pedir_tamano(siguiente_unidad, cantidad_total)

        total = sum(c["precio"] for c in carrito)
        _guardar_sesion(
            telefono, sesion["nombre"], "esperando_confirmacion",
            cantidad_total=cantidad_total, unidad_actual=siguiente_unidad, carrito=carrito,
        )
        resumen = _resumen_carrito(carrito)
        return _botones(
            f"Tu pedido:\n{resumen}\n\n*Total: S/ {total:.2f}*\n\n¿Confirmas este pedido?",
            [("confirmar_pedido", "✅ Confirmar"), ("cancelar_pedido", "❌ Cancelar")],
        )

    # --- Paso 3: confirmación antes de pedir ubicación ---
    if paso == "esperando_confirmacion":
        opciones_confirmacion = [("confirmar_pedido", "✅ Confirmar"), ("cancelar_pedido", "❌ Cancelar")]
        decision = _detectar_confirmacion(_resolver_por_posicion(texto, opciones_confirmacion))
        if decision == "cancelar":
            _borrar_sesion(telefono)
            return _texto("Tu pedido fue cancelado. Escribe cualquier mensaje para empezar de nuevo. 🙂")
        if decision == "confirmar":
            _guardar_sesion(
                telefono, sesion["nombre"], "esperando_ubicacion",
                cantidad_total=sesion["cantidad_total"], carrito=sesion["carrito"],
            )
            return _texto(
                f"Ahora comparte tu *ubicación* (GPS) para verificar si estás dentro de nuestra zona "
                f"de reparto ({RADIO_MAXIMO_KM} km)."
            )
        resumen = _resumen_carrito(sesion["carrito"])
        total = sum(c["precio"] for c in sesion["carrito"])
        return _botones(
            f"Tu pedido:\n{resumen}\n\n*Total: S/ {total:.2f}*\n\n¿Confirmas este pedido?",
            [("confirmar_pedido", "✅ Confirmar"), ("cancelar_pedido", "❌ Cancelar")],
        )

    # --- Paso 4: ubicación ---
    if paso == "esperando_ubicacion":
        if tipo != "ubicacion" or lat is None or lon is None:
            return _texto("Por favor comparte tu ubicación (GPS) usando el botón/función de ubicación de WhatsApp.")

        distancia = calcular_distancia(LAT_LOCAL, LON_LOCAL, lat, lon)
        if distancia > RADIO_MAXIMO_KM:
            _borrar_sesion(telefono)
            return _texto(
                f"Lo sentimos, tu ubicación está a {distancia:.1f} km y nuestro rango máximo de delivery "
                f"es de {RADIO_MAXIMO_KM} km. 😔\n\n"
                f"Mientras tanto puedes recogerlo tú mismo en nuestro local: {DIRECCION_LOCAL}. "
                f"¡Pronto llegaremos más lejos! Escribe cualquier mensaje si deseas intentar con otra dirección."
            )

        carrito = sesion["carrito"]
        cliente_nombre = sesion["nombre"] or nombre
        total_unidades = sesion["cantidad_total"]
        monto_total = sum(c["precio"] for c in carrito)
        detalle = _detalle_agrupado(carrito)

        with _lock:
            cursor.execute(
                """
                INSERT INTO pedidos
                    (telefono, cliente_nombre, detalle, total_unidades, monto_total, latitud, longitud, distancia_km)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                """,
                (telefono, cliente_nombre, detalle, total_unidades, monto_total, lat, lon, round(distancia, 2)),
            )
            conn.commit()

        _borrar_sesion(telefono)
        return _texto(
            f"¡Pedido confirmado! ✅\n"
            f"{detalle}\n"
            f"Total: S/ {monto_total:.2f}\n"
            f"Distancia: {distancia:.1f} km\n\n"
            f"Gracias, {cliente_nombre}. Tu pedido está en camino. 🍌\n\n"
            f"_Si quieres agregar algo más, escríbelo y con gusto te ayudamos con un pedido nuevo._"
        )

    # Estado desconocido: reiniciar
    _borrar_sesion(telefono)
    return _pedir_cantidad(nombre)


def obtener_pedido(pedido_id: int):
    with _lock:
        cursor.execute(
            "SELECT id, telefono, cliente_nombre, detalle, monto_total, estado FROM pedidos WHERE id = %s",
            (pedido_id,),
        )
        f = cursor.fetchone()
    if not f:
        return None
    return {
        "id": f[0], "telefono": f[1], "cliente_nombre": f[2],
        "detalle": f[3], "monto_total": f[4], "estado": f[5],
    }


def marcar_entregado(pedido_id: int):
    """Antes main.py tocaba bl.cursor directamente para esto. Se movió acá
    para que toda la base de datos se maneje desde un solo lugar."""
    with _lock:
        cursor.execute("UPDATE pedidos SET estado = 'ENTREGADO' WHERE id = %s", (pedido_id,))
        conn.commit()


def guardar_suscripcion_push(endpoint: str, p256dh: str, auth: str):
    with _lock:
        cursor.execute(
            """
            INSERT INTO suscripciones_push (endpoint, p256dh, auth)
            VALUES (%s, %s, %s)
            ON CONFLICT (endpoint) DO UPDATE SET
                p256dh = excluded.p256dh,
                auth = excluded.auth
            """,
            (endpoint, p256dh, auth),
        )
        conn.commit()


def eliminar_suscripcion_push(endpoint: str):
    with _lock:
        cursor.execute("DELETE FROM suscripciones_push WHERE endpoint = %s", (endpoint,))
        conn.commit()


def mensaje_entregado(cliente_nombre: str) -> str:
    """Mensaje de agradecimiento que se envía al cliente cuando el
    repartidor marca su pedido como 'Entregado'."""
    saludo = f"¡Gracias, {cliente_nombre}, " if cliente_nombre else "¡Gracias "
    return (
        f"{saludo}por disfrutar los ricos sabores de Maduritos Asados! 🍌🔥\n\n"
        f"No olvides que estamos todos los días en {DIRECCION_LOCAL}.\n\n"
        f"¡Pronto llegaremos más lejos y te esperamos de nuevo! 😊"
    )


# ---------------------------------------------------------------------------
# HISTORIAL Y REPORTES (idea: historial por cliente + reportes por periodo)
# ---------------------------------------------------------------------------
def historial_cliente(telefono: str):
    with _lock:
        cursor.execute(
            """
            SELECT id, detalle, total_unidades, monto_total, distancia_km, estado, fecha
            FROM pedidos WHERE telefono = %s ORDER BY fecha DESC
            """,
            (telefono,),
        )
        filas = cursor.fetchall()
    pedidos = [
        {
            "id": f[0], "detalle": f[1], "total_unidades": f[2], "monto_total": f[3],
            "distancia_km": f[4], "estado": f[5], "fecha": f[6],
        }
        for f in filas
    ]
    total_gastado = sum(p["monto_total"] for p in pedidos if p["estado"] == "ENTREGADO")
    return {"telefono": telefono, "pedidos": pedidos, "total_gastado_soles": round(total_gastado, 2)}


def reporte_por_periodo(periodo: str = "dia"):
    condiciones = {
        "dia": "date(fecha) = CURRENT_DATE",
        "semana": "date(fecha) >= CURRENT_DATE - INTERVAL '6 days'",
        "mes": "to_char(fecha, 'YYYY-MM') = to_char(CURRENT_DATE, 'YYYY-MM')",
    }
    condicion = condiciones.get(periodo, condiciones["dia"])
    with _lock:
        cursor.execute(
            f"""
            SELECT id, total_unidades, monto_total, detalle
            FROM pedidos WHERE estado = 'ENTREGADO' AND {condicion}
            """
        )
        filas = cursor.fetchall()

    total_pedidos = len(filas)
    total_maduros = sum(f[1] for f in filas)
    total_dinero = sum(f[2] for f in filas)

    # Nuevo: Contador de insumos para el ranking
    ranking = {"Queso": 0, "Maní": 0, "Chicharrón": 0}
    for f in filas:
        detalle = (f[3] or "").lower()
        for linea in detalle.split('\n'):
            if 'x' in linea:
                try:
                    cant = int(linea.split('x')[0].strip())
                    if 'queso' in linea: ranking['Queso'] += cant
                    if 'maní' in linea or 'mani' in linea: ranking['Maní'] += cant
                    if 'chicharrón' in linea or 'chicharron' in linea: ranking['Chicharrón'] += cant
                except:
                    pass

    return {
        "periodo": periodo,
        "pedidos_entregados": total_pedidos,
        "maduritos_vendidos": total_maduros,
        "total_recaudado_soles": round(total_dinero, 2),
        "ranking": ranking
    }


def pedidos_detallados_periodo(periodo: str = "dia"):
    condiciones = {
        "dia": "date(fecha) = CURRENT_DATE",
        "semana": "date(fecha) >= CURRENT_DATE - INTERVAL '6 days'",
        "mes": "to_char(fecha, 'YYYY-MM') = to_char(CURRENT_DATE, 'YYYY-MM')",
    }
    condicion = condiciones.get(periodo, condiciones["dia"])
    with _lock:
        cursor.execute(
            f"""
            SELECT id, cliente_nombre, telefono, detalle, total_unidades, monto_total, fecha
            FROM pedidos WHERE estado = 'ENTREGADO' AND {condicion}
            ORDER BY fecha ASC
            """
        )
        filas = cursor.fetchall()
        
    resultado = []
    for f in filas:
        nombre_db = f[1] or "Sin nombre"
        tel_db = f[2] or ""
        
        # Limpiamos si viene con el formato "Nombre | Celular"
        if " | " in nombre_db:
            partes = nombre_db.split(" | ")
            nombre_limpio = partes[0]
            tel_limpio = partes[1].strip()
        else:
            nombre_limpio = nombre_db
            tel_limpio = tel_db.split("@")[0] # Quitamos el @lid
            
        # Formateamos la fecha para quitar los milisegundos feos
        fecha_limpia = str(f[6])[:16] if f[6] else ""
            
        resultado.append({
            "id": f[0], "cliente_nombre": nombre_limpio, "telefono": tel_limpio,
            "detalle": f[3], "total_unidades": f[4], "monto_total": f[5], "fecha": fecha_limpia,
        })
    return resultado


def generar_excel_reporte(periodo: str = "dia") -> bytes:
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    # Importamos las herramientas para tablas oficiales
    from openpyxl.worksheet.table import Table, TableStyleInfo

    pedidos = pedidos_detallados_periodo(periodo)
    resumen = reporte_por_periodo(periodo)
    etiquetas = {"dia": "Hoy", "semana": "Últimos 7 días", "mes": "Este mes"}

    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de Caja"

    titulo_font = Font(name="Arial", size=14, bold=True, color="FFFFFF")
    titulo_fill = PatternFill("solid", fgColor="00A884")
    encabezado_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    encabezado_fill = PatternFill("solid", fgColor="2E7D32")

    ws.merge_cells("A1:F1")
    ws["A1"] = f"Maduritos Asados — Cierre de Caja ({etiquetas.get(periodo, periodo)})"
    ws["A1"].font = titulo_font
    ws["A1"].fill = titulo_fill
    ws["A1"].alignment = Alignment(horizontal="center")

    ws["A3"] = "Total recaudado (S/)"
    ws["B3"] = resumen["total_recaudado_soles"]
    ws["A4"] = "Maduritos vendidos"
    ws["B4"] = resumen["maduritos_vendidos"]
    ws["A5"] = "Pedidos entregados"
    ws["B5"] = resumen["pedidos_entregados"]
    for fila in (3, 4, 5):
        ws[f"A{fila}"].font = Font(name="Arial", bold=True)

    encabezados = ["ID", "Cliente", "Teléfono", "Detalle", "Maduritos", "Monto", "Fecha"]
    fila_encabezado = 7
    for col, titulo in enumerate(encabezados, start=1):
        celda = ws.cell(row=fila_encabezado, column=col, value=titulo)
        celda.font = encabezado_font
        celda.fill = encabezado_fill

    for i, p in enumerate(pedidos, start=fila_encabezado + 1):
        ws.cell(row=i, column=1, value=p["id"])
        ws.cell(row=i, column=2, value=p["cliente_nombre"])
        ws.cell(row=i, column=3, value=p["telefono"])
        celda_detalle = ws.cell(row=i, column=4, value=p["detalle"])
        celda_detalle.alignment = Alignment(wrap_text=True)
        ws.cell(row=i, column=5, value=p["total_unidades"])
        ws.cell(row=i, column=6, value=p["monto_total"])
        ws.cell(row=i, column=7, value=p["fecha"])

    anchos = [6, 18, 16, 40, 11, 12, 20]
    for col, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[chr(64 + col)].width = ancho

    # Aplicamos el formato de Tabla Inteligente de Excel
    if pedidos:
        tab = Table(displayName="DatosCaja", ref=f"A7:G{fila_encabezado + len(pedidos)}")
        estilo_tabla = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        tab.tableStyleInfo = estilo_tabla
        ws.add_table(tab)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


def generar_pdf_reporte(periodo: str = "dia") -> bytes:
    from io import BytesIO
    # 1. Agregamos "landscape" a las importaciones
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet

    pedidos = pedidos_detallados_periodo(periodo)
    resumen = reporte_por_periodo(periodo)
    etiquetas = {"dia": "Hoy", "semana": "Últimos 7 días", "mes": "Este mes"}

    buffer = BytesIO()
    
    # 2. Volteamos la hoja a horizontal (landscape) y reducimos los márgenes de los lados
    doc = SimpleDocTemplate(
        buffer, 
        pagesize=landscape(letter), 
        topMargin=1.5 * cm, bottomMargin=1.5 * cm,
        leftMargin=1.0 * cm, rightMargin=1.0 * cm
    )
    estilos = getSampleStyleSheet()
    elementos = []

    elementos.append(Paragraph(
        f"<b>Maduritos Asados</b> — Cierre de Caja ({etiquetas.get(periodo, periodo)})",
        estilos["Title"],
    ))
    elementos.append(Spacer(1, 12))
    elementos.append(Paragraph(
        f"Total recaudado: <b>S/ {resumen['total_recaudado_soles']:.2f}</b> &nbsp;&nbsp; "
        f"Maduritos vendidos: <b>{resumen['maduritos_vendidos']}</b> &nbsp;&nbsp; "
        f"Pedidos entregados: <b>{resumen['pedidos_entregados']}</b>",
        estilos["Normal"],
    ))
    elementos.append(Spacer(1, 16))

    datos = [["ID", "Cliente", "Teléfono", "Detalle", "Uds", "Monto", "Fecha"]]
    for p in pedidos:
        # Mantenemos el salto de línea HTML que agregamos antes
        detalle_html = p["detalle"].replace("\n", "<br/>")
        datos.append([
            str(p["id"]), p["cliente_nombre"], p["telefono"],
            Paragraph(detalle_html, estilos["Normal"]),
            str(p["total_unidades"]), f"S/ {p['monto_total']:.2f}", p["fecha"],
        ])

    # 3. Ampliamos las columnas para usar todo el ancho de la hoja horizontal
    # La columna del detalle ahora tiene 11.5 cm (antes tenía 6.5 cm)
    tabla = Table(datos, colWidths=[1.2 * cm, 3.5 * cm, 2.5 * cm, 11.5 * cm, 1.2 * cm, 2.2 * cm, 3.5 * cm], repeatRows=1)
    
    tabla.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2E7D32")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F5F5")]),
    ]))
    elementos.append(tabla)

    doc.build(elementos)
    buffer.seek(0)
    return buffer.read()


def pedidos_pendientes():
    with _lock:
        cursor.execute(
            """
            SELECT id, cliente_nombre, telefono, detalle, monto_total, latitud, longitud, distancia_km, fecha
            FROM pedidos WHERE estado = 'PENDIENTE' ORDER BY id ASC
            """
        )
        filas = cursor.fetchall()
    return [
        {
            "id": f[0], "cliente_nombre": f[1] or "Sin nombre", "telefono": f[2], "detalle": f[3],
            "monto_total": f[4], "latitud": f[5], "longitud": f[6], "distancia_km": f[7], "fecha": f[8],
        }
        for f in filas
    ]
